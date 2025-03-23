import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
import os

class SistemaInventario:
    def __init__(self, archivo="inventario.xlsx"):
        if not os.path.exists(archivo):
            self.crear_archivo_inicial(archivo)
        self.archivo = archivo
        self.workbook = openpyxl.load_workbook(archivo)
        self.sheet = self.workbook["Inventario"]

    def crear_archivo_inicial(self, archivo):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Inventario"
        headers = ["ID", "Producto", "Cantidad", "Precio Unitario", "Última Modificación"]
        sheet.append(headers)
        wb.save(archivo)

    def agregar_producto(self):
        producto = input("Ingrese el nombre del producto: ").strip()
        cantidad = int(input("Ingrese la cantidad inicial: "))
        precio = float(input("Ingrese el precio unitario: "))
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        nuevo_id = self.sheet.max_row
        self.sheet.append([nuevo_id, producto, cantidad, precio, fecha])
        self.workbook.save(self.archivo)
        print(f"Producto {producto} agregado correctamente.")

    def actualizar_existencias(self):
        producto = input("Ingrese el nombre del producto: ").strip()
        nueva_cantidad = int(input("Ingrese la nueva cantidad: "))
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for row in range(2, self.sheet.max_row + 1):
            if self.sheet.cell(row=row, column=2).value == producto:
                self.sheet.cell(row=row, column=3).value = nueva_cantidad
                self.sheet.cell(row=row, column=5).value = fecha
                self.workbook.save(self.archivo)
                print(f"Existencias de {producto} actualizadas a {nueva_cantidad}.")
                return
        print("Producto no encontrado.")

    def registrar_venta(self):
        producto = input("Ingrese el nombre del producto vendido: ").strip()
        cantidad_vendida = int(input("Ingrese la cantidad vendida: "))
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for row in range(2, self.sheet.max_row + 1):
            if self.sheet.cell(row=row, column=2).value == producto:
                cantidad_actual = int(self.sheet.cell(row=row, column=3).value)
                if cantidad_actual >= cantidad_vendida:
                    self.sheet.cell(row=row, column=3).value = cantidad_actual - cantidad_vendida
                    self.sheet.cell(row=row, column=5).value = fecha
                    self.workbook.save(self.archivo)
                    print(f"Venta registrada: {cantidad_vendida} unidades de {producto}.")
                else:
                    print("No hay suficiente stock disponible.")
                return
        print("Producto no encontrado.")

    def generar_reporte(self):
        df = pd.read_excel(self.archivo)
        print("\nInventario Actual:")
        print(df.to_string(index=False))

    def visualizar_estadisticas(self):
        df = pd.read_excel(self.archivo)
        if df.empty:
            print("No hay datos en el inventario.")
            return
        df.plot(x="Producto", y="Cantidad", kind="bar", title="Stock de Productos", figsize=(10, 5))
        plt.xticks(rotation=45)
        plt.show()


def menu_principal():
    sistema = SistemaInventario()
    while True:
        print("\n===== SISTEMA DE INVENTARIO =====")
        print("1. Agregar producto")
        print("2. Actualizar existencias")
        print("3. Registrar venta")
        print("4. Generar reporte")
        print("5. Visualizar estadísticas")
        print("6. Salir")
        opcion = input("Seleccione una opción: ")
        if opcion == "1":
            sistema.agregar_producto()
        elif opcion == "2":
            sistema.actualizar_existencias()
        elif opcion == "3":
            sistema.registrar_venta()
        elif opcion == "4":
            sistema.generar_reporte()
        elif opcion == "5":
            sistema.visualizar_estadisticas()
        elif opcion == "6":
            print("¡Hasta luego!")
            break
        else:
            print("Opción no válida. Intente de nuevo.")

if __name__ == "__main__":
    menu_principal()